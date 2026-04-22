"""Serial bridge web UI for the Pico DC motor controller.

This server runs on your desktop (Windows/macOS/Linux) and talks to the
MicroPython ``main.py`` running on the Pico Robotics board via USB serial. It
serves the ``web_ui_dc.html`` page so you can drive motors with sliders and
keyboard controls, mirroring the Raspberry Pi web interface but without any
Blinka dependencies.

Launch example::

    python web_control_server.py --serial COM12 --http-port 8765
"""

from __future__ import annotations

import argparse
import csv
import threading
import time
from datetime import datetime
from pathlib import Path

import math

from flask import Flask, jsonify, request, send_from_directory
import serial

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

app = Flask(__name__)

SERIAL_LOCK = threading.Lock()
STATE_LOCK = threading.Lock()
SERIAL_CONN: serial.Serial | None = None

HTML_PAGE = (Path(__file__).parent / "web_ui_dc.html").read_text(encoding="utf-8")

STATUS_CACHE = {
    "motors": [{"motor": idx, "throttle": 0.0} for idx in range(1, 5)],
    "increment": 0.10,
    "updated": 0.0,
}
ANGLE_CACHE = {
    "available": False,
    "values": [],
    "updated": 0.0,
    "error": "Waiting for Pico data",
}
RAW_CACHE = {
    "available": False,
    "values": [],
    "updated": 0.0,
    "error": "Waiting for Pico potentiometers",
}
TELEMETRY_CACHE = {
    "available": False,
    "latest": None,
    "updated": 0.0,
    "error": "Waiting for receiver telemetry",
}
CALIBRATION_STATE = {
    "running": False,
    "motor": None,
    "stage": "unsupported",
    "progress": 0.0,
    "message": "Calibration handled on Pico",
    "error": None,
    "updated": time.time(),
}
AUTO_CACHE = {
    "enabled": False,
    "state": "off",
    "motor": 0,
    "dir": 0,
    "rssi": None,
    "delta": None,
    "updated": 0.0,
}
LOCK_CACHE = {
    "has_fix": False,
    "lock_announced": False,
    "event_seq": 0,
    "locked_at": 0.0,
    "block_until": 0.0,
    "updated": 0.0,
}
MANUAL_OVERRIDE_BLOCK_SECONDS = 2.0

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
DATA_STALE_SECONDS = 3.0
LOG_LOCK = threading.Lock()
LOG_COLUMNS = [
    "timestamp_iso",
    "timestamp_unix",
    "counter",
    "fix",
    "gps_time",
    "temp",
    "pressure",
    "baro",
    "humidity",
    "lat",
    "lon",
    "alt",
    "speed",
    "rx_rssi",
    "azimuth_deg",
    "elevation_deg",
    "azimuth_raw",
    "elevation_raw",
    "azimuth_ohms",
    "elevation_ohms",
    "auto_state",
    "auto_motor",
    "auto_dir",
    "auto_delta",
]
XLSX_PATH = DATA_DIR / "telemetry_log.xlsx"
CSV_PATH = DATA_DIR / "telemetry_log.csv"


def _sanitize_for_log(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned.lower() in {"", "nan", "nl", "nofix", "no data", "none", "--"}:
            return ""
        return cleaned
    return value


def _cache_is_fresh(updated_ts: float, now_ts: float) -> bool:
    if not isinstance(updated_ts, (int, float)):
        return False
    if updated_ts <= 0:
        return False
    return (now_ts - float(updated_ts)) <= DATA_STALE_SECONDS


def _value_by_name(entries, name, key):
    if not isinstance(entries, list):
        return ""
    for entry in entries:
        if str(entry.get("name", "")).lower() != name:
            continue
        return _sanitize_for_log(entry.get(key))
    return ""


_LOG_WORKBOOK = None
_LOG_SHEET = None


def _init_log_targets():
    global _LOG_WORKBOOK, _LOG_SHEET
    if Workbook is not None and load_workbook is not None:
        try:
            if XLSX_PATH.exists():
                _LOG_WORKBOOK = load_workbook(XLSX_PATH)
                _LOG_SHEET = _LOG_WORKBOOK.active
                if _LOG_SHEET.max_row < 1:
                    _LOG_SHEET.append(LOG_COLUMNS)
            else:
                _LOG_WORKBOOK = Workbook()
                _LOG_SHEET = _LOG_WORKBOOK.active
                _LOG_SHEET.title = "telemetry"
                _LOG_SHEET.append(LOG_COLUMNS)
                _LOG_WORKBOOK.save(XLSX_PATH)
            print(f"Data logging enabled: {XLSX_PATH}")
            return
        except Exception as exc:
            _LOG_WORKBOOK = None
            _LOG_SHEET = None
            print(f"[WARN] XLSX logger unavailable ({exc}); falling back to CSV")

    try:
        csv_needs_header = not CSV_PATH.exists()
        with CSV_PATH.open("a", newline="", encoding="utf-8") as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=LOG_COLUMNS)
            if csv_needs_header:
                writer.writeheader()
        print(f"Data logging enabled (CSV fallback): {CSV_PATH}")
    except Exception as exc:
        print(f"[WARN] Data logger disabled: {exc}")


def _write_log_row(row: dict) -> None:
    sanitized = {key: _sanitize_for_log(row.get(key)) for key in LOG_COLUMNS}
    with LOG_LOCK:
        if _LOG_WORKBOOK is not None and _LOG_SHEET is not None:
            try:
                _LOG_SHEET.append([sanitized.get(col, "") for col in LOG_COLUMNS])
                _LOG_WORKBOOK.save(XLSX_PATH)
                return
            except Exception as exc:
                print(f"[WARN] XLSX write failed ({exc}); switching to CSV fallback")

        try:
            with CSV_PATH.open("a", newline="", encoding="utf-8") as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=LOG_COLUMNS)
                writer.writerow(sanitized)
        except Exception as exc:
            print(f"[WARN] CSV write failed: {exc}")


def _log_telemetry_snapshot(payload: dict) -> None:
    now_ts = time.time()
    now_iso = datetime.fromtimestamp(now_ts).isoformat(timespec="seconds")

    with STATE_LOCK:
        angle_values = [dict(entry) for entry in ANGLE_CACHE.get("values", [])]
        angle_updated = ANGLE_CACHE.get("updated", 0.0)
        raw_values = [dict(entry) for entry in RAW_CACHE.get("values", [])]
        raw_updated = RAW_CACHE.get("updated", 0.0)
        auto_state = dict(AUTO_CACHE)
        auto_updated = AUTO_CACHE.get("updated", 0.0)

    use_angles = _cache_is_fresh(angle_updated, now_ts)
    use_raw = _cache_is_fresh(raw_updated, now_ts)
    use_auto = _cache_is_fresh(auto_updated, now_ts)

    row = {
        "timestamp_iso": now_iso,
        "timestamp_unix": now_ts,
        "counter": payload.get("c"),
        "fix": payload.get("fix"),
        "gps_time": payload.get("gps_time"),
        "temp": payload.get("temp"),
        "pressure": payload.get("pressure"),
        "baro": payload.get("baro"),
        "humidity": payload.get("humidity"),
        "lat": payload.get("lat"),
        "lon": payload.get("lon"),
        "alt": payload.get("alt"),
        "speed": payload.get("speed"),
        "rx_rssi": payload.get("rx_rssi"),
        "azimuth_deg": _value_by_name(angle_values, "azimuth", "degrees") if use_angles else "",
        "elevation_deg": _value_by_name(angle_values, "elevation", "degrees") if use_angles else "",
        "azimuth_raw": _value_by_name(raw_values, "azimuth", "raw") if use_raw else "",
        "elevation_raw": _value_by_name(raw_values, "elevation", "raw") if use_raw else "",
        "azimuth_ohms": _value_by_name(raw_values, "azimuth", "ohms") if use_raw else "",
        "elevation_ohms": _value_by_name(raw_values, "elevation", "ohms") if use_raw else "",
        "auto_state": auto_state.get("state") if use_auto else "",
        "auto_motor": auto_state.get("motor") if use_auto else "",
        "auto_dir": auto_state.get("dir") if use_auto else "",
        "auto_delta": auto_state.get("delta") if use_auto else "",
    }
    _write_log_row(row)


def ensure_serial_open() -> serial.Serial:
    conn = SERIAL_CONN
    if conn is None or not conn.is_open:
        raise RuntimeError("Serial port is not open")
    return conn


def send_line(text: str) -> None:
    # Leading space keeps Pico's hotkey handler from consuming the first letter
    sanitized = text.strip()
    if not sanitized:
        return
    payload = (" " + sanitized + "\n").encode("utf-8")
    with SERIAL_LOCK:
        conn = ensure_serial_open()
        conn.write(payload)
        conn.flush()


def _update_throttles(line: str) -> None:
    parts = line.split(",")[1:]
    motors = []
    for idx, token in enumerate(parts, start=1):
        token = token.strip()
        throttle = None
        if token.upper() != "REL":
            try:
                throttle = float(token)
            except Exception:
                print(f"[WARN] Unable to parse throttle token '{token}' for motor {idx}")
                return
        motors.append({"motor": idx, "throttle": throttle})
    if motors:
        with STATE_LOCK:
            STATUS_CACHE["motors"] = motors
            STATUS_CACHE["updated"] = time.time()


def _update_angles(line: str) -> None:
    entries = []
    for token in line.split(",")[1:]:
        if "=" not in token:
            continue
        name, value = token.split("=", 1)
        value = value.strip()
        if value.lower().endswith("deg"):
            value = value[:-3]
        try:
            degrees = float(value)
        except Exception:
            continue
        entries.append({"name": name.strip(), "degrees": degrees})
    with STATE_LOCK:
        ANGLE_CACHE.update(
            {
                "available": bool(entries),
                "values": entries,
                "updated": time.time(),
                "error": None if entries else "No angle data",
            }
        )


def _update_raw(line: str) -> None:
    entries = []
    for token in line.split(",")[1:]:
        if "=" not in token:
            continue
        name, value = token.split("=", 1)
        value = value.strip()
        if ":" in value:
            parts = value.split(":", 1)
            raw = parts[0]
            ohms = parts[1]
        else:
            raw = value
            ohms = "nan"
        try:
            raw_int = int(raw)
        except Exception:
            raw_int = None
        try:
            ohms_val = float(ohms)
        except Exception:
            ohms_val = None
        if ohms_val is not None and not math.isfinite(ohms_val):
            ohms_val = None
        entries.append({"name": name.strip(), "raw": raw_int, "ohms": ohms_val})
    with STATE_LOCK:
        RAW_CACHE.update(
            {
                "available": bool(entries),
                "values": entries,
                "updated": time.time(),
                "error": None if entries else "No raw data",
            }
        )


def _parse_tel_value(key: str, value: str):
    if value.strip().lower() in {"nan", "nofix", "nl", "", "no data", "none", "--"}:
        return None
    if key in {"c", "fix"}:
        try:
            return int(value)
        except Exception:
            return value
    if key in {"temp", "pressure", "baro", "humidity", "lat", "lon", "alt", "speed", "rx_rssi"}:
        try:
            number = float(value)
        except Exception:
            return None
        if math.isfinite(number):
            return number
        return None
    return value


def _update_telemetry(line: str) -> None:
    payload = {}
    for token in line.split(",")[1:]:
        if "=" not in token:
            continue
        key, value = token.split("=", 1)
        payload[key.strip()] = _parse_tel_value(key.strip(), value.strip())
    now = time.time()
    with STATE_LOCK:
        TELEMETRY_CACHE.update(
            {
                "available": bool(payload),
                "latest": payload if payload else None,
                "updated": now,
                "error": None if payload else "Empty telemetry payload",
            }
        )
        if payload:
            has_fix = payload.get("fix") == 1
            prev_fix = bool(LOCK_CACHE.get("has_fix"))
            LOCK_CACHE["has_fix"] = bool(has_fix)
            if has_fix and not prev_fix and not LOCK_CACHE.get("lock_announced"):
                LOCK_CACHE["lock_announced"] = True
                LOCK_CACHE["event_seq"] = int(LOCK_CACHE.get("event_seq", 0)) + 1
                LOCK_CACHE["locked_at"] = now
                LOCK_CACHE["block_until"] = now + MANUAL_OVERRIDE_BLOCK_SECONDS
                LOCK_CACHE["updated"] = now
    if payload:
        _log_telemetry_snapshot(payload)


def _update_auto(line: str) -> None:
    payload = {}
    for token in line.split(",")[1:]:
        if "=" not in token:
            continue
        key, value = token.split("=", 1)
        payload[key.strip()] = value.strip()

    with STATE_LOCK:
        enabled_raw = payload.get("enabled")
        AUTO_CACHE["enabled"] = str(enabled_raw) in {"1", "true", "True", "ON", "on"}
        AUTO_CACHE["state"] = payload.get("state", AUTO_CACHE["state"])
        try:
            AUTO_CACHE["motor"] = int(payload.get("motor", AUTO_CACHE["motor"]))
        except Exception:
            pass
        try:
            AUTO_CACHE["dir"] = int(payload.get("dir", AUTO_CACHE["dir"]))
        except Exception:
            pass
        try:
            parsed_rssi = float(payload.get("rssi"))
            AUTO_CACHE["rssi"] = parsed_rssi if math.isfinite(parsed_rssi) else None
        except Exception:
            AUTO_CACHE["rssi"] = None
        try:
            parsed_delta = float(payload.get("delta"))
            AUTO_CACHE["delta"] = parsed_delta if math.isfinite(parsed_delta) else None
        except Exception:
            AUTO_CACHE["delta"] = None
        AUTO_CACHE["updated"] = time.time()


def _handle_serial_line(decoded: str) -> None:
    print(f"[PICO] {decoded}", end="")
    stripped = decoded.strip()
    if not stripped:
        return

    frames = (
        ("THROT,", _update_throttles),
        ("ANGLES,", _update_angles),
        ("POTRAW,", _update_raw),
        ("TEL,", _update_telemetry),
        ("AUTO,", _update_auto),
    )

    # 1) Fast path: regular clean lines.
    for prefix, handler in frames:
        if stripped.startswith(prefix):
            handler(stripped)
            return

    # 2) Prompt-prefixed path: "motor> TEL,..." or ">>> THROT,...".
    for prompt_sep in ("motor>", ">>>", "...", "] "):
        if prompt_sep not in stripped:
            continue
        tail = stripped.split(prompt_sep, 1)[1].strip()
        for prefix, handler in frames:
            if tail.startswith(prefix):
                handler(tail)
                return


def serial_reader() -> None:
    while True:
        try:
            conn = ensure_serial_open()
        except Exception:
            time.sleep(0.2)
            continue
        try:
            line = conn.readline()
        except Exception:
            time.sleep(0.2)
            continue
        if not line:
            continue
        try:
            decoded = line.decode("utf-8", errors="replace")
        except Exception:
            decoded = repr(line)
        _handle_serial_line(decoded)


def compose_status() -> dict:
    with STATE_LOCK:
        _normalize_lock_state_locked()
        motors = [dict(entry) for entry in STATUS_CACHE["motors"]]
        increment = STATUS_CACHE["increment"]
        updated = STATUS_CACHE["updated"]
        angle_copy = {
            "available": ANGLE_CACHE["available"],
            "values": [dict(entry) for entry in ANGLE_CACHE["values"]],
            "updated": ANGLE_CACHE["updated"],
            "error": ANGLE_CACHE.get("error"),
        }
        raw_copy = {
            "available": RAW_CACHE["available"],
            "values": [dict(entry) for entry in RAW_CACHE["values"]],
            "updated": RAW_CACHE["updated"],
            "error": RAW_CACHE.get("error"),
        }
        telemetry_copy = {
            "available": TELEMETRY_CACHE["available"],
            "latest": dict(TELEMETRY_CACHE["latest"]) if isinstance(TELEMETRY_CACHE.get("latest"), dict) else None,
            "updated": TELEMETRY_CACHE["updated"],
            "error": TELEMETRY_CACHE.get("error"),
        }
        calibration_copy = dict(CALIBRATION_STATE)
        auto_copy = dict(AUTO_CACHE)
        lock_copy = dict(LOCK_CACHE)
        block_until = float(lock_copy.get("block_until", 0.0) or 0.0)
        lock_copy["remaining"] = max(0.0, block_until - time.time())
    return {
        "motors": motors,
        "increment": increment,
        "updated": updated,
        "angles": angle_copy,
        "raw": raw_copy,
        "telemetry": telemetry_copy,
        "calibration": calibration_copy,
        "auto": auto_copy,
        "lock": lock_copy,
    }


@app.route("/")
def index():
    return HTML_PAGE


@app.route("/soundeffects/<path:filename>")
def soundeffects(filename: str):
    sound_dir = Path(__file__).parent / "soundeffects"
    return send_from_directory(str(sound_dir), filename)


def _normalize_lock_state_locked(now: float | None = None) -> None:
    now_ts = time.time() if now is None else float(now)
    raw_until = LOCK_CACHE.get("block_until", 0.0)
    try:
        block_until = float(raw_until or 0.0)
    except Exception:
        block_until = 0.0
    if not math.isfinite(block_until) or block_until <= now_ts:
        LOCK_CACHE["block_until"] = 0.0


def _manual_override_locked() -> tuple[bool, float]:
    with STATE_LOCK:
        _normalize_lock_state_locked()
        block_until = float(LOCK_CACHE.get("block_until", 0.0) or 0.0)
    remaining = block_until - time.time()
    if remaining > 0:
        return True, remaining
    return False, 0.0


def _is_manual_text_command(text: str) -> bool:
    token = (text.strip().split() or [""])[0].lower()
    return token in {"set", "inc", "dec", "nudge", "w", "s", "i", "k"}


def _handle_calibration_commands(cmd: str):
    return jsonify({"error": "Calibration is not supported via Pico web bridge."}), 400


@app.route("/command", methods=["POST"])
def command():
    data = request.get_json(silent=True) or {}
    text = (data.get("text") or "").strip()
    if text:
        is_locked, remaining = _manual_override_locked()
        if is_locked and _is_manual_text_command(text):
            return jsonify({"error": f"Manual override locked for {remaining:.1f}s after GPS lock"}), 429
        try:
            send_line(text)
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        return jsonify({"ok": True, "message": f"Sent '{text}'", "status": compose_status()})

    cmd = (data.get("cmd") or "").strip().lower()
    if not cmd:
        return jsonify({"error": "Missing command"}), 400
    if cmd in {"calibrate", "calibration_stop"}:
        return _handle_calibration_commands(cmd)

    try:
        _dispatch_command(cmd, data)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    return jsonify({"ok": True, "message": "Command sent", "status": compose_status()})


def _require_motor(data: dict) -> int:
    motor = data.get("motor")
    if motor is None:
        raise ValueError("Motor is required")
    try:
        return int(motor)
    except Exception as exc:
        raise ValueError("Motor must be an integer") from exc


def _dispatch_command(cmd: str, data: dict) -> None:
    if cmd in {"nudge", "set_throttle"}:
        is_locked, remaining = _manual_override_locked()
        if is_locked:
            raise ValueError(f"Manual override locked for {remaining:.1f}s after GPS lock")

    if cmd == "nudge":
        motor = _require_motor(data)
        try:
            direction_input = float(data.get("direction", 1))
        except Exception:
            raise ValueError("Direction must be numeric")
        direction = 1 if direction_input >= 0 else -1
        amount = data.get("amount")
        suffix = ""
        if amount is not None:
            try:
                amt = abs(float(amount))
                suffix = f" {amt:.2f}"
            except Exception:
                raise ValueError("Amount must be numeric")
        base = "inc" if direction >= 0 else "dec"
        send_line(f"{base} {motor}{suffix}")
        return
    if cmd == "set_throttle":
        motor = _require_motor(data)
        value = data.get("value")
        if value is None:
            raise ValueError("Throttle value is required")
        try:
            throttle = float(value)
        except Exception as exc:
            raise ValueError("Throttle must be numeric") from exc
        send_line(f"set {motor} {throttle:.3f}")
        return
    if cmd == "release":
        if data.get("motor") is None:
            send_line("release")
        else:
            send_line(f"release {_require_motor(data)}")
        return
    if cmd == "stop":
        motor = _require_motor(data)
        send_line(f"stop {motor}")
        return
    if cmd == "stopall":
        send_line("stopall")
        return
    if cmd == "increment":
        amount = data.get("amount")
        if amount is None:
            raise ValueError("Amount is required")
        try:
            value = float(amount)
        except Exception:
            raise ValueError("Amount must be numeric")
        send_line(f"increment {value:.2f}")
        with STATE_LOCK:
            STATUS_CACHE["increment"] = max(0.0, min(1.0, value))
        return
    if cmd == "auto_track":
        enabled = data.get("enabled")
        is_enabled = bool(enabled)
        base_lat = data.get("base_lat")
        base_lon = data.get("base_lon")
        base_alt = data.get("base_alt")

        has_base = base_lat is not None and base_lon is not None
        if has_base:
            try:
                lat = float(base_lat)
                lon = float(base_lon)
                alt = 0.0 if base_alt is None else float(base_alt)
            except Exception:
                raise ValueError("Base lat/lon/alt must be numeric")
            send_line(f"auto base {lat:.6f} {lon:.6f} {alt:.1f}")
        elif not is_enabled:
            # When disabling auto without valid base coordinates from UI,
            # clear stale base target on the Pico side.
            send_line("auto clearbase")

        send_line("auto on" if is_enabled else "auto off")
        with STATE_LOCK:
            AUTO_CACHE["enabled"] = is_enabled
            AUTO_CACHE["updated"] = time.time()
        return
    raise ValueError(f"Unknown command: {cmd}")


@app.route("/status")
def status():
    return jsonify(compose_status())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Web UI bridge for Pico DC motors")
    parser.add_argument("--serial", required=True, help="Serial/COM port of the Pico (e.g. COM5 or /dev/ttyACM0)")
    parser.add_argument("--baud", type=int, default=115200, help="Baud rate (default: 115200)")
    parser.add_argument("--host", default="127.0.0.1", help="HTTP bind address (default: 127.0.0.1)")
    parser.add_argument("--http-port", type=int, default=8765, help="HTTP port for the web server (default: 8765)")
    return parser.parse_args()


def main() -> None:
    global SERIAL_CONN
    args = parse_args()
    _init_log_targets()
    print(f"Opening serial port {args.serial} @ {args.baud}...")
    SERIAL_CONN = serial.Serial(args.serial, args.baud, timeout=0.1)
    print("Serial connection established. Launching reader thread...")
    threading.Thread(target=serial_reader, daemon=True).start()
    print(f"Serving DC web UI at http://{args.host}:{args.http_port}")
    print("Press Ctrl+C to stop.")
    app.run(host=args.host, port=args.http_port, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()
